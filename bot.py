import os
import io
import time
import logging
import requests
import pandas as pd
from openpyxl import load_workbook
from telegram import Update
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters,
)
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# ── Настройки ──────────────────────────────────────────────────────────────────

BOT_TOKEN      = os.getenv("TELEGRAM_BOT_TOKEN")
ONEDRIVE_SHARE_URL = os.getenv("ONEDRIVE_SHARE_URL")

if not BOT_TOKEN:
    raise ValueError("TELEGRAM_BOT_TOKEN environment variable is required")
if not ONEDRIVE_SHARE_URL:
    raise ValueError("ONEDRIVE_SHARE_URL environment variable is required")

CACHE_TTL = 60  # секунд — как часто обновлять файл из OneDrive
_cache: dict = {"result": None, "ts": 0.0}

# Внутренние имена колонок после нормализации
COL_TRACKING = "_tracking"
COL_CLIENT   = "_client"
COL_DESC     = "_desc"
COL_SENT     = "_sent"
COL_METHOD   = "_method"
COL_WEIGHT   = "_weight"
COL_PRICE    = "_price"
COL_NOTES    = "_notes"

# ── Логирование ────────────────────────────────────────────────────────────────

logging.basicConfig(
    format="%(asctime)s │ %(levelname)s │ %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ── Microsoft Graph API ────────────────────────────────────────────────────────

def download_excel_bytes() -> bytes:
    """Скачивает Excel файл из OneDrive sharing-ссылки.
    Открывает страницу просмотра и извлекает прямую ссылку на скачивание из HTML/JSON.
    """
    import re

    headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}

    # Шаг 1: получаем страницу просмотра OneDrive
    r = requests.get(ONEDRIVE_SHARE_URL, headers=headers, timeout=30, allow_redirects=True)
    html = r.text
    logger.info("Страница просмотра: статус %s, размер %d", r.status_code, len(html))

    # Шаг 2: ищем прямую download ссылку внутри HTML (OneDrive прячет её в JSON)
    patterns = [
        r'"downloadUrl"\s*:\s*"([^"]+)"',
        r'"url"\s*:\s*"(https://[^"]*\.xlsx[^"]*)"',
        r'downloadUrl["\s:]+["\']?(https://[^\s"\'<>]+)',
        r'"FileGetUrl"\s*:\s*"([^"]+)"',
        r'sj\.u\(["\']([^"\']*download[^"\']*)["\']',
    ]

    download_url = None
    for pattern in patterns:
        match = re.search(pattern, html)
        if match:
            download_url = match.group(1).replace("\\u0026", "&").replace("\\/", "/")
            logger.info("Найдена download ссылка через паттерн: %s", pattern)
            break

    if download_url:
        r2 = requests.get(download_url, headers=headers, timeout=30, allow_redirects=True)
        ct = r2.headers.get("Content-Type", "")
        logger.info("Скачивание: статус %s, Content-Type: %s, размер: %d", r2.status_code, ct, len(r2.content))
        if r2.status_code == 200 and len(r2.content) > 5000:
            return r2.content

    # Шаг 3: если не нашли — логируем кусок HTML для диагностики
    logger.error("Download URL не найден. Кусок HTML:\n%s", html[:3000])
    raise RuntimeError("Не удалось найти прямую ссылку на скачивание в OneDrive")

# ── Зелёные строки (не прошли контроль) ───────────────────────────────────────

def is_row_green(ws, row_idx: int) -> bool:
    """True если строка закрашена зелёным (RGB, theme или indexed цвет)."""
    green = 0
    total = 0
    for cell in ws[row_idx]:
        if cell.value is None:
            continue
        total += 1
        fill = cell.fill
        if not fill or fill.fill_type != "solid":
            continue
        try:
            color = fill.fgColor
            if color.type == "rgb":
                rgb = str(color.rgb)
                if len(rgb) == 8 and rgb not in ("FFFFFFFF", "00000000", "FF000000", "00FFFFFF"):
                    r_val = int(rgb[2:4], 16)
                    g_val = int(rgb[4:6], 16)
                    b_val = int(rgb[6:8], 16)
                    if g_val > r_val and g_val > b_val and g_val > 60:
                        green += 1
            elif color.type == "theme":
                # Зелёные тона в стандартных Excel темах: индексы 6, 9 (Accent 3, 6)
                if color.theme in (6, 9):
                    green += 1
            elif color.type == "indexed":
                # Стандартные зелёные indexed цвета в Excel
                if color.indexed in (4, 10, 17, 50, 57):
                    green += 1
        except Exception:
            pass
    return total > 0 and (green / total) > 0.3

# ── Нормализация листов ────────────────────────────────────────────────────────

def normalize_sheet(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame | None:
    """Переименовывает колонки реального Excel под внутренние имена.
    Поддерживает AVIA (SF трек в Unnamed: 4) и CARGO (трек в колонке с ТРЕК-КОД).
    """
    # Стрипаем все имена колонок
    df.columns = [str(c).strip() for c in df.columns]

    # Ищем колонку с трек-кодом
    tracking_col = None
    for col in df.columns:
        if "ТРЕК" in col.upper() or "三方单号" in col:
            tracking_col = col
            break

    # Для AVIA листов трек-коды в Unnamed: 4
    if tracking_col is None and "AVIA" in sheet_name.upper():
        if "Unnamed: 4" in df.columns:
            tracking_col = "Unnamed: 4"

    if tracking_col is None:
        return None

    # Метод доставки из названия листа
    if "AVIA" in sheet_name.upper():
        method = "авиа"
    elif "CARGO" in sheet_name.upper():
        method = "наземная"
    else:
        method = ""

    # Строим маппинг: ищем каждую колонку по ключевым словам
    rename = {tracking_col: COL_TRACKING}
    df[COL_METHOD] = method

    for col in df.columns:
        u = col.upper()
        if col == tracking_col:
            continue
        if ("ДАТА" in u or "ОТПРАВКИ" in u or "发货日期" in col) and COL_SENT not in rename.values():
            rename[col] = COL_SENT
        elif ("НАЗВАНИЕ" in u or "品名" in col) and COL_DESC not in rename.values():
            rename[col] = COL_DESC
        elif ("ПОЛУЧАТЕЛ" in u or "ИМЯ" in u or "КЛИЕНТ" in u or "收件人" in col) and COL_CLIENT not in rename.values():
            # Приоритет: ПОЛУЧАТЕЛЬ / ИМЯ / КЛИЕНТ — не путаем с номером заказа
            rename[col] = COL_CLIENT
        elif ("ПУНКТ" in u or "目的地" in col) and COL_CLIENT not in rename.values():
            rename[col] = COL_CLIENT
        elif ("ВЕС" in u or "重量" in col) and "货代" not in col and COL_WEIGHT not in rename.values():
            rename[col] = COL_WEIGHT
        elif ("ПОЛУЧИТЬ" in u or "运费" in col) and COL_PRICE not in rename.values():
            rename[col] = COL_PRICE
        elif ("КОММЕНТ" in u or "备注" in col) and COL_NOTES not in rename.values():
            rename[col] = COL_NOTES

    df = df.rename(columns=rename)

    # Нормализуем трек-коды: убираем пробелы, переводим в верхний регистр.
    # Числовые коды pandas читает как float (1234567.0) — убираем дробную часть.
    def _norm_track(val) -> str:
        s = str(val).strip()
        # "1234567890.0" → "1234567890"
        if s.endswith(".0") and s[:-2].isdigit():
            s = s[:-2]
        return s.upper()

    df[COL_TRACKING] = df[COL_TRACKING].apply(_norm_track)

    # Убираем пустые строки (где трек-код пустой или nan)
    df = df[~df[COL_TRACKING].isin(["", "NAN", "NONE"])]

    return df


# ── Работа с Excel ─────────────────────────────────────────────────────────────

def load_orders():
    """Скачивает Excel из OneDrive, читает ВСЕ листы, находит зелёные строки.
    Возвращает (DataFrame со всеми заказами, set трек-кодов зелёных строк).
    Результат кешируется на CACHE_TTL секунд.
    """
    now = time.time()
    if _cache["result"] and now - _cache["ts"] < CACHE_TTL:
        return _cache["result"]

    excel_bytes = download_excel_bytes()
    buf = io.BytesIO(excel_bytes)

    # ── Читаем все листы через pandas ──
    xl = pd.ExcelFile(buf)
    logger.info("Листы в файле: %s", xl.sheet_names)
    frames = []

    for sheet in xl.sheet_names:
        try:
            # Первая строка — китайский заголовок, вторая строка — реальные колонки
            df_raw = pd.read_excel(xl, sheet_name=sheet, header=1, dtype=str)
            df_norm = normalize_sheet(df_raw, sheet)
            if df_norm is not None and not df_norm.empty:
                df_norm["_sheet"] = sheet
                frames.append(df_norm)
                logger.info("✅ Лист «%s»: %d строк", sheet, len(df_norm))
            else:
                logger.warning("Лист «%s» — трек-колонка не найдена или пуста", sheet)
        except Exception as e:
            logger.warning("Лист «%s» пропущен: %s", sheet, e)

    if not frames:
        logger.error("Ни один лист не содержит трек-коды")
        result = (pd.DataFrame(), set())
        _cache["result"] = result
        _cache["ts"] = now
        return result

    df_all = pd.concat(frames, ignore_index=True)

    # ── Зелёные строки через openpyxl ──
    buf.seek(0)
    wb = load_workbook(buf, data_only=True)
    green_tracks: set[str] = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        is_avia = "AVIA" in sheet_name.upper()

        # Ищем заголовочную строку: ищем ячейку с "ТРЕК" или "三方单号"
        header_row_idx = None
        track_col_idx  = None

        for row in ws.iter_rows(min_row=1, max_row=5):
            for cell in row:
                val = str(cell.value or "").strip()
                if "ТРЕК" in val.upper() or "三方单号" in val:
                    header_row_idx = cell.row
                    track_col_idx  = cell.column
                    break
            if header_row_idx:
                break

        # Для AVIA: трек в столбце 5 (Unnamed: 4 → индекс 4 → столбец 5)
        if header_row_idx is None and is_avia:
            header_row_idx = 2   # строка заголовка
            track_col_idx  = 5   # столбец 5 (E)

        if header_row_idx is None:
            continue

        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            track_val = ws.cell(row=row_idx, column=track_col_idx).value
            if track_val and is_row_green(ws, row_idx):
                green_tracks.add(str(track_val).strip().upper())

    logger.info("Зелёных строк (не прошли контроль): %d", len(green_tracks))
    result = (df_all, green_tracks)
    _cache["result"] = result
    _cache["ts"] = now
    return result


def find_order(code: str):
    """Ищет заказ по трек-коду.
    Возвращает (order_dict | None, status: str)
    status может быть:
      'not_found'          — трек-код не найден
      'ok'                 — посылка в пути, всё хорошо
      'transferred_to_cargo' — не прошла авиа-контроль, переведена на карго
      'detained'           — задержана на авиа-складе, ещё не в карго
    """
    df, green_tracks = load_orders()
    # Нормализуем введённый код: верхний регистр, убираем пробелы и ".0"
    code = code.strip().upper()
    if code.endswith(".0") and code[:-2].isdigit():
        code = code[:-2]
    matches = df[df[COL_TRACKING] == code]
    # Запасной поиск: если точное совпадение не найдено — ищем подстроку
    if matches.empty:
        matches = df[df[COL_TRACKING].str.contains(code, na=False, regex=False)]
    if matches.empty:
        return None, "not_found"

    is_green = code in green_tracks
    cargo_matches = matches[matches["_sheet"].str.upper().str.contains("CARGO")]
    avia_matches  = matches[matches["_sheet"].str.upper().str.contains("AVIA")]

    if is_green:
        if not cargo_matches.empty:
            # Зелёная в AVIA → переведена на карго (и там тоже зелёная)
            return cargo_matches.iloc[0].to_dict(), "transferred_to_cargo"
        else:
            # Только в AVIA и зелёная — ещё не перенесена
            row = avia_matches.iloc[0] if not avia_matches.empty else matches.iloc[0]
            return row.to_dict(), "detained"
    else:
        # Всё нормально: приоритет карго над авиа
        result_row = cargo_matches.iloc[0] if not cargo_matches.empty else matches.iloc[0]
        return result_row.to_dict(), "ok"


def fmt_date(value) -> str:
    if pd.isna(value) or str(value).strip() in ("", "nan", "NaT"):
        return "—"
    try:
        return pd.Timestamp(value).strftime("%d.%m.%Y")
    except Exception:
        return str(value)


def calc_arrival(sent_value, method: str) -> str:
    """Рассчитывает примерный диапазон дат прибытия.
    Авиа: +3-4 дня, Наземная: +7-12 дней от даты отправки.
    """
    try:
        if pd.isna(sent_value) or str(sent_value).strip() in ("", "nan", "NaT"):
            return "—"
        sent_ts = pd.Timestamp(sent_value)
        m = str(method).lower()
        if "авиа" in m or "air" in m:
            d1 = (sent_ts + pd.Timedelta(days=3)).strftime("%d.%m.%Y")
            d2 = (sent_ts + pd.Timedelta(days=4)).strftime("%d.%m.%Y")
        else:  # наземная доставка
            d1 = (sent_ts + pd.Timedelta(days=7)).strftime("%d.%m.%Y")
            d2 = (sent_ts + pd.Timedelta(days=12)).strftime("%d.%m.%Y")
        return f"{d1} — {d2}"
    except Exception:
        return "—"


def fmt_method(value) -> str:
    s = str(value).strip().lower()
    if "авиа" in s or "air" in s:
        return "✈️ авиа"
    if "наземн" in s or "ground" in s or "land" in s:
        return "🚚 наземная"
    return f"📦 {value}"


def get_val(order: dict, col: str) -> str:
    v = str(order.get(col, "")).strip()
    return "—" if v in ("", "nan", "None", "NaT") else v


def build_reply(order: dict, header: str | None = None) -> str:
    raw_method = get_val(order, COL_METHOD)
    method     = fmt_method(raw_method)
    sent       = fmt_date(order.get(COL_SENT))
    arrival    = calc_arrival(order.get(COL_SENT), raw_method)
    client     = get_val(order, COL_CLIENT)
    desc       = get_val(order, COL_DESC)
    notes      = get_val(order, COL_NOTES)

    try:
        price_str = f"💰 *Стоимость:* ${float(get_val(order, COL_PRICE)):,.2f}"
    except Exception:
        price_str = f"💰 *Стоимость:* {get_val(order, COL_PRICE)}"

    try:
        weight_str = f"⚖️ *Вес:* {float(get_val(order, COL_WEIGHT))} кг"
    except Exception:
        weight_str = f"⚖️ *Вес:* {get_val(order, COL_WEIGHT)}"

    if header:
        title_line = header
    else:
        title_line = f"📦 *Ваш товар отправлен* {sent}"

    lines = [
        title_line,
        f"Способ доставки: {method}",
        "",
        "────────────────────",
        f"🔖 *Трек-код:* `{get_val(order, COL_TRACKING)}`",
        f"👤 *Получатель:* {client}",
        f"📝 *Товар:* {desc}",
        price_str,
        weight_str,
        f"📅 *Дата отправки:* {sent}",
        f"🕐 *Примерная дата прибытия:* {arrival}",
    ]
    if notes != "—":
        lines.append(f"📌 *Примечание:* {notes}")

    lines += ["────────────────────", "_По вопросам обращайтесь в нашу службу поддержки._"]
    return "\n".join(lines)


def build_reply_transferred(order: dict) -> str:
    """Сообщение для посылки, которая не прошла авиа и переведена на карго."""
    sent  = fmt_date(order.get(COL_SENT))
    header = (
        "⚠️ *Ваша посылка не прошла авиа-контроль*\n\n"
        "Отправление было возвращено на склад и переведено на *наземную доставку (карго)*.\n\n"
        f"🚚 Новая дата отправки: *{sent}*"
    )
    return build_reply(order, header=header)

# ── Обработчики ────────────────────────────────────────────────────────────────

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text(
        "👋 *Здравствуйте!*\n"
        "Вас приветствует Telegram-бот *6016.kz*.\n\n"
        "Чем могу быть полезен?",
        parse_mode="Markdown",
    )
    await update.message.reply_text(
        "📦 Пожалуйста, отправьте *трек-код* товара, чтобы получить информацию о доставке.\n\n"
        "Примеры трек-кодов:\n"
        "`SF1234567891011`\n"
        "`JDK1234567890`\n"
        "`1234567890`\n\n"
        "Поддерживаются все форматы курьерских служб.",
        parse_mode="Markdown",
    )


async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text(
        "📦 *Как узнать статус доставки:*\n\n"
        "Просто отправьте ваш *трек-код* товара.\n\n"
        "Примеры:\n"
        "`SF1234567891011`\n"
        "`JDK1234567890`\n"
        "`1234567890`\n\n"
        "Трек-код можно найти в чеке или подтверждении заказа.\n\n"
        "🔄 Для принудительного обновления данных используйте /refresh",
        parse_mode="Markdown",
    )


async def refresh_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Принудительно сбрасывает кеш и перезагружает данные из Excel."""
    _cache["result"] = None
    _cache["ts"] = 0.0
    await update.message.reply_text(
        "🔄 Кеш сброшен. Данные будут обновлены при следующем запросе.",
        parse_mode="Markdown",
    )


async def track(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_input = update.message.text.strip()
    user = update.effective_user
    logger.info("Запрос от %s (%s): %r", user.id, user.first_name, user_input)

    await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")

    try:
        order, status = find_order(user_input)
    except Exception as e:
        logger.error("Ошибка при загрузке данных: %s", e)
        await update.message.reply_text(
            "⚠️ Не удалось загрузить данные. Попробуйте позже.",
            parse_mode="Markdown",
        )
        return

    if order is None:
        # Товар не найден вообще
        await update.message.reply_text(
            "⚠️ *Упс…*\n\n"
            "Похоже, товар ещё не был отправлен или не поступил на наш склад.\n\n"
            "🔍 Попробуйте проверить статус позже.",
            parse_mode="Markdown",
        )
    elif status == "transferred_to_cargo":
        # Не прошла авиа-контроль → переведена на карго
        await update.message.reply_text(
            "✅ *Товар найден!*\n\nСейчас проверяю информацию… ⏳",
            parse_mode="Markdown",
        )
        await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
        await update.message.reply_text(build_reply_transferred(order), parse_mode="Markdown")
    elif status == "detained":
        # Задержана на авиа-складе, в карго ещё не появилась
        await update.message.reply_text(
            "✅ *Товар найден!*\n\nСейчас проверяю информацию… ⏳",
            parse_mode="Markdown",
        )
        await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
        await update.message.reply_text(
            "⚠️ *Ваша посылка не прошла авиа-контроль.*\n\n"
            "Отправление находится на складе и готовится к переводу на *наземную доставку (карго)*.\n\n"
            "📞 По вопросам обращайтесь в нашу службу поддержки.",
            parse_mode="Markdown",
        )
    else:
        # Товар найден, всё нормально (status == 'ok')
        await update.message.reply_text(
            "✅ *Товар найден!*\n\nСейчас проверяю информацию о доставке… ⏳",
            parse_mode="Markdown",
        )
        await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
        await update.message.reply_text(build_reply(order), parse_mode="Markdown")

# ── Запуск ─────────────────────────────────────────────────────────────────────

def main() -> None:
    logger.info("Бот запускается…")
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start",   start))
    app.add_handler(CommandHandler("help",    help_cmd))
    app.add_handler(CommandHandler("refresh", refresh_cmd))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, track))
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
